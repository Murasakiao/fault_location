from openpyxl import load_workbook # type: ignore
import numpy as np # type: ignore
import math
import cmath

def extract_data_from_sheet(file_path, sheet_name, min, max):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []
    all_data = []
    for row in sheet.iter_rows(min_row=min, max_row=max, values_only=True):
        data.append(row)
    for row in data:
        all_data.extend(row)
    return all_data

def three_phase_fault(fault_currents, fault_voltages, system_data):
    # assign base values
    Vf = 1
    base_mva = system_data[5]
    base_kv1 = system_data[1]
    base_kv2 = system_data[7]
    base_kv3 = system_data[15]
    # assign actual values
    act_mva_gen = system_data[0]
    act_kv_gen = system_data[1]
    act_mva_T1 = system_data[5]
    act_kv1_T1 = system_data[6]
    act_kv2_T1 = system_data[7]
    act_mva_T2 = system_data[13]
    act_kv1_T2 = system_data[14]
    act_kv2_T2 = system_data[15]
    act_mw_load = system_data[19]
    act_mvar_load = system_data[20]
    act_R_fault = system_data[21]
    act_X_fault = system_data[22]
    act_z_ground = system_data[23]
    # assign per unit values 
    # generator
    XG1 = system_data[2] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG2 = system_data[3] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG0 = system_data[4] * (base_mva / act_mva_gen) * (act_kv2_T1 / base_kv2)**2
    # transformer 1
    XT11 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT12 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT10 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    # transmission line
    ZLine = complex(system_data[11], system_data[12]) / (base_kv2**2 / base_mva)
    # transformer 2
    XT21 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT22 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT20 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    # load
    ZLoad = Vf**2 * (base_mva / act_mw_load)
    # load current
    ILoad = act_mw_load / base_mva
    # assign phase values
    Ia = cmath.rect(fault_currents[0][0], math.radians(fault_currents[0][1]))
    Ib = cmath.rect(fault_currents[1][0], math.radians(fault_currents[1][1]))
    Ic = cmath.rect(fault_currents[2][0], math.radians(fault_currents[2][1]))
    Va = cmath.rect(fault_voltages[0][0], math.radians(fault_voltages[0][1]))
    Vb = cmath.rect(fault_voltages[1][0], math.radians(fault_voltages[1][1]))
    Vc = cmath.rect(fault_voltages[2][0], math.radians(fault_voltages[2][1]))
    a = complex(-0.5, math.sqrt(3)/2)
    A = np.array([[1, 1, 1], [1, a, a**2], [1, a**2, a]])
    Iabc = np.array([Ia, Ib, Ic])
    Vabc = np.array([Va, Vb, Vc])
    # sequence components
    Ia012 = 1/3 * (np.dot(A, Iabc))
    Va012 = 1/3 * (np.dot(A, Vabc))
    Ia0 = Ia012[0]
    Ia1 = Ia012[1]
    Ia2 = Ia012[2]
    Va0 = Va012[0]
    Va1 = Va012[1]
    Va2 = Va012[2]

    def at_bus_1(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv1**2 / base_mva)
        Z1 = (Vf / Ia1) - ZFault
        Z2 = 0
        Z0 = 0
        Zx = complex(0, XG1) 
        Zy = complex(0, XT11) + ZLine + complex(0, XT21) + ZLoad
        Zeq = Zx * Zy / (Zx + Zy)
        return np.round(Zeq, 4) == np.round(Z1, 4)

    def at_bus_2(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        Z1 = (Vf / Ia1) - ZFault
        Z2 = 0
        Z0 = 0
        Zx = complex(0, XG1) + complex(0, XT11) 
        Zy = ZLine + complex(0, XT21) + ZLoad
        Zeq = Zx * Zy / (Zx + Zy)
        return np.round(Zeq, 4) == np.round(Z1, 4)

    def at_bus_3(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        Z1 = (Vf / Ia1) - ZFault
        Z2 = 0
        Z0 = 0
        Zx = complex(0, XG1) + complex(0, XT11) + ZLine
        Zy = complex(0, XT21) + ZLoad
        Zeq = Zx * Zy / (Zx + Zy)
        return np.round(Zeq, 4) == np.round(Z1, 4)

    def at_bus_4(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv3**2 / base_mva)
        Z1 = (Vf / Ia1) - ZFault
        Z2 = 0
        Z0 = 0
        Zx = complex(0, XG1) + complex(0, XT11) + ZLine + complex(0, XT21)
        Zy = ZLoad
        Zeq = Zx * Zy / (Zx + Zy)
        return np.round(Zeq, 4) == np.round(Z1, 4)
    
    # Check which case the parameter Z1 fits into
    if at_bus_1(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 1"
    elif at_bus_2(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 2"
    elif at_bus_3(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 3"
    elif at_bus_4(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 4"
    else:
        return "Unknown case"

def SLG_fault(fault_currents, fault_voltages, system_data):
    # assign base values
    Vf = 1
    base_mva = system_data[5]
    base_kv1 = system_data[1]
    base_kv2 = system_data[7]
    base_kv3 = system_data[15]
    # assign actual values
    act_mva_gen = system_data[0]
    act_kv_gen = system_data[1]
    act_mva_T1 = system_data[5]
    act_kv1_T1 = system_data[6]
    act_kv2_T1 = system_data[7]
    act_mva_T2 = system_data[13]
    act_kv1_T2 = system_data[14]
    act_kv2_T2 = system_data[15]
    act_mw_load = system_data[19]
    act_mvar_load = system_data[20]
    act_R_fault = system_data[21]
    act_X_fault = system_data[22]
    act_z_ground = system_data[23]
    # assign per unit values 
    # generator
    XG1 = system_data[2] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG2 = system_data[3] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG0 = system_data[4] * (base_mva / act_mva_gen) * (act_kv2_T1 / base_kv2)**2
    # transformer 1
    XT11 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT12 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT10 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    # transmission line
    ZLine = complex(system_data[11], system_data[12]) / (base_kv2**2 / base_mva)
    # transformer 2
    XT21 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT22 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT20 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    # load
    ZLoad = Vf**2 * (base_mva / act_mw_load)
    # fault resistance
    ZFault = act_R_fault / (base_kv2**2 / base_mva)
    # load current
    ILoad = act_mw_load / base_mva
    # assign phase values
    Ia = cmath.rect(fault_currents[0][0], math.radians(fault_currents[0][1]))
    Ib = cmath.rect(fault_currents[1][0], math.radians(fault_currents[1][1]))
    Ic = cmath.rect(fault_currents[2][0], math.radians(fault_currents[2][1]))
    Va = cmath.rect(fault_voltages[0][0], math.radians(fault_voltages[0][1]))
    Vb = cmath.rect(fault_voltages[1][0], math.radians(fault_voltages[1][1]))
    Vc = cmath.rect(fault_voltages[2][0], math.radians(fault_voltages[2][1]))
    a = complex(-0.5, math.sqrt(3)/2)
    A = np.array([[1, 1, 1], [1, a, a**2], [1, a**2, a]])
    Iabc = np.array([Ia, Ib, Ic])
    Vabc = np.array([Va, Vb, Vc])
    # sequence components
    Ia012 = 1/3 * (np.dot(A, Iabc))
    Va012 = 1/3 * (np.dot(A, Vabc))
    Ia0 = Ia012[0]
    Ia1 = Ia012[1]
    Ia2 = Ia012[2]
    Va0 = Va012[0]
    Va1 = Va012[1]
    Va2 = Va012[2]

    r = [abs(Ia012[2]), math.degrees(cmath.phase(Ia012[2]))]
    t = [abs(Va012[0]), math.degrees(cmath.phase(Va012[0]))]
    
    def at_bus_1(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv1**2 / base_mva)
        Z1 = (Vf / Ia1) - ZFault
        Z2 = - (Va2 / Ia2)
        Z0 = 0
        Zx = complex(0, XG1) 
        Zy = complex(0, XT11) + ZLine + complex(0, XT21) + ZLoad
        Zeq = Zx * Zy / (Zx + Zy)
        return np.round(Zeq, 4) == np.round(Z1, 4)

    def at_bus_2(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        Z1 = (Vf / Ia1) - ZFault
        Z2 = 0
        Z0 = 0
        Zx = complex(0, XG1) + complex(0, XT11) 
        Zy = ZLine + complex(0, XT21) + ZLoad
        Zeq = Zx * Zy / (Zx + Zy)
        return np.round(Zeq, 4) == np.round(Z1, 4)

    def at_bus_3(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        Z1 = (Vf - Va1) / Ia1
        Z2 = -(Va2 / Ia2)
        Z0 = -(Va0 / Ia0)
        Zx = complex(0, XG1) + complex(0, XT11) + ZLine
        Zy = complex(0, XT21) + ZLoad
        Zeq = Zx * Zy / (Zx + Zy)
        return np.round(Zeq, 4) == np.round(Z1, 4)

    def at_bus_4(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv3**2 / base_mva)
        Z1 = (Vf / Ia1) - ZFault
        Z2 = 0
        Z0 = 0
        Zx = complex(0, XG1) + complex(0, XT11) + ZLine + complex(0, XT21)
        Zy = ZLoad
        Zeq = Zx * Zy / (Zx + Zy)
        return np.round(Zeq, 4) == np.round(Z1, 4)
    
    # Check which case the parameter Z1 fits into
    if at_bus_1(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 1"
    elif at_bus_2(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 2"
    elif at_bus_3(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 3"
    elif at_bus_4(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 4"
    else:
        return "Unknown case"

def LL_fault(fault_currents, fault_voltages, system_data):
    # assign base values
    Vf = 1
    base_mva = system_data[5]
    base_kv1 = system_data[1]
    base_kv2 = system_data[7]
    base_kv3 = system_data[15]
    # assign actual values
    act_mva_gen = system_data[0]
    act_kv_gen = system_data[1]
    act_mva_T1 = system_data[5]
    act_kv1_T1 = system_data[6]
    act_kv2_T1 = system_data[7]
    act_mva_T2 = system_data[13]
    act_kv1_T2 = system_data[14]
    act_kv2_T2 = system_data[15]
    act_mw_load = system_data[19]
    act_mvar_load = system_data[20]
    act_R_fault = system_data[21]
    act_X_fault = system_data[22]
    act_z_ground = system_data[23]
    # assign per unit values 
    # generator
    XG1 = system_data[2] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG2 = system_data[3] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG0 = system_data[4] * (base_mva / act_mva_gen) * (act_kv2_T1 / base_kv2)**2
    # transformer 1
    XT11 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT12 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT10 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    # transmission line
    ZLine = complex(system_data[11], system_data[12]) / (base_kv2**2 / base_mva)
    # transformer 2
    XT21 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT22 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT20 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    # load
    ZLoad = Vf**2 * (base_mva / act_mw_load)
    # fault resistance
    ZFault = act_R_fault / (base_kv2**2 / base_mva)
    # load current
    ILoad = act_mw_load / base_mva
    # assign phase values
    Ia = cmath.rect(fault_currents[0][0], math.radians(fault_currents[0][1]))
    Ib = cmath.rect(fault_currents[1][0], math.radians(fault_currents[1][1]))
    Ic = cmath.rect(fault_currents[2][0], math.radians(fault_currents[2][1]))
    Va = cmath.rect(fault_voltages[0][0], math.radians(fault_voltages[0][1]))
    Vb = cmath.rect(fault_voltages[1][0], math.radians(fault_voltages[1][1]))
    Vc = cmath.rect(fault_voltages[2][0], math.radians(fault_voltages[2][1]))
    a = complex(-0.5, math.sqrt(3)/2)
    A = np.array([[1, 1, 1], [1, a, a**2], [1, a**2, a]])
    Iabc = np.array([Ia, Ib, Ic])
    Vabc = np.array([Va, Vb, Vc])
    # sequence components
    Ia012 = 1/3 * (np.dot(A, Iabc))
    Va012 = 1/3 * (np.dot(A, Vabc))
    Ia0 = Ia012[0]
    Ia1 = Ia012[1]
    Ia2 = Ia012[2]
    Va0 = Va012[0]
    Va1 = Va012[1]
    Va2 = Va012[2]

    def at_bus_1(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv1**2 / base_mva)
        Z1 = (Vf - Va1) / Ia1
        Z2 = -(Va2 / Ia2)
        Z0 = -(Va0 / Ia0)
        Zx = complex(0, XG1) 
        Zy = complex(0, XT11) + ZLine + complex(0, XT21) + ZLoad
        Zeq = Zx * Zy / (Zx + Zy)

        return np.round(Zeq, 4) == np.round(Z1, 4)

    def at_bus_2(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        Z1 = (Vf - Va1) / Ia1
        Z2 = -(Va2 / Ia2)
        Z0 = -(Va0 / Ia0)
        Zx = complex(0, XG1) + complex(0, XT11) 
        Zy = ZLine + complex(0, XT21) + ZLoad
        Zeq = Zx * Zy / (Zx + Zy)
        return np.round(Zeq, 4) == np.round(Z1, 4)

    def at_bus_3(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        Z1 = (Vf - Va1) / Ia1
        Z2 = -(Va2 / Ia2)
        Z0 = -(Va0 / Ia0)
        Zx = complex(0, XG1) + complex(0, XT11) + ZLine
        Zy = complex(0, XT21) + ZLoad
        Zeq = Zx * Zy / (Zx + Zy)
        return np.round(Zeq, 4) == np.round(Z1, 4)

    def at_bus_4(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv3**2 / base_mva)
        Z1 = (Vf - Va1) / Ia1
        Z2 = -(Va2 / Ia2)
        Z0 = -(Va0 / Ia0)
        Zx = complex(0, XG1) + complex(0, XT11) + ZLine + complex(0, XT21)
        Zy = ZLoad
        Zeq = Zx * Zy / (Zx + Zy)
        return np.round(Zeq, 4) == np.round(Z1, 4)
    
    # Check which case the parameter Z1 fits into
    if at_bus_1(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 1"
    elif at_bus_2(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 2"
    elif at_bus_3(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 3"
    elif at_bus_4(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 4"
    else:
        return "Unknown case"

def DLG_fault(fault_currents, fault_voltages, system_data):
    # assign base values
    Vf = 1
    base_mva = system_data[5]
    base_kv1 = system_data[1]
    base_kv2 = system_data[7]
    base_kv3 = system_data[15]
    # assign actual values
    act_mva_gen = system_data[0]
    act_kv_gen = system_data[1]
    act_mva_T1 = system_data[5]
    act_kv1_T1 = system_data[6]
    act_kv2_T1 = system_data[7]
    act_mva_T2 = system_data[13]
    act_kv1_T2 = system_data[14]
    act_kv2_T2 = system_data[15]
    act_mw_load = system_data[19]
    act_mvar_load = system_data[20]
    act_R_fault = system_data[21]
    act_X_fault = system_data[22]
    act_z_ground = system_data[23]
    # assign per unit values 
    # generator
    XG1 = system_data[2] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG2 = system_data[3] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG0 = system_data[4] * (base_mva / act_mva_gen) * (act_kv2_T1 / base_kv2)**2
    # transformer 1
    XT11 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT12 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT10 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    # transmission line
    ZLine = complex(system_data[11], system_data[12]) / (base_kv2**2 / base_mva)
    # transformer 2
    XT21 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT22 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT20 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    # load
    ZLoad = Vf**2 * (base_mva / act_mw_load)
    # fault resistance
    ZFault = act_R_fault / (base_kv2**2 / base_mva)
    # load current
    ILoad = act_mw_load / base_mva
    # assign phase values
    Ia = cmath.rect(fault_currents[0][0], math.radians(fault_currents[0][1]))
    Ib = cmath.rect(fault_currents[1][0], math.radians(fault_currents[1][1]))
    Ic = cmath.rect(fault_currents[2][0], math.radians(fault_currents[2][1]))
    Va = cmath.rect(fault_voltages[0][0], math.radians(fault_voltages[0][1]))
    Vb = cmath.rect(fault_voltages[1][0], math.radians(fault_voltages[1][1]))
    Vc = cmath.rect(fault_voltages[2][0], math.radians(fault_voltages[2][1]))
    a = complex(-0.5, math.sqrt(3)/2)
    A = np.array([[1, 1, 1], [1, a, a**2], [1, a**2, a]])
    Iabc = np.array([Ia, Ib, Ic])
    Vabc = np.array([Va, Vb, Vc])
    # sequence components
    Ia012 = 1/3 * (np.dot(A, Iabc))
    Va012 = 1/3 * (np.dot(A, Vabc))
    Ia0 = Ia012[0]
    Ia1 = Ia012[1]
    Ia2 = Ia012[2]
    Va0 = Va012[0]
    Va1 = Va012[1]
    Va2 = Va012[2]
    Z1 = (Vf - Va1) / Ia1
    Z2 = -(Va2 / Ia2)
    Z0 = -(Va0 / Ia0)
    
    def at_bus_1(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv1**2 / base_mva)
        Z1 = (Vf - Va1) / Ia1
        Z2 = -(Va2 / Ia2)
        Z0 = -(Va0 / Ia0)
        Zx = complex(0, XG1) 
        Zy = complex(0, XT11) + ZLine + complex(0, XT21) + ZLoad
        Zeq = Zx * Zy / (Zx + Zy)
        return np.round(Zeq, 4) == np.round(Z1, 4)

    def at_bus_2(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        Z1 = (Vf - Va1) / Ia1
        Z2 = -(Va2 / Ia2)
        Z0 = -(Va0 / Ia0)
        Zx = complex(0, XG1) + complex(0, XT11) 
        Zy = ZLine + complex(0, XT21) + ZLoad
        Zeq = Zx * Zy / (Zx + Zy)
        return np.round(Zeq, 4) == np.round(Z1, 4)

    def at_bus_3(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        Z1 = (Vf - Va1) / Ia1
        Z2 = -(Va2 / Ia2)
        Z0 = -(Va0 / Ia0)
        Zx = complex(0, XG1) + complex(0, XT11) + ZLine
        Zy = complex(0, XT21) + ZLoad
        Zeq = Zx * Zy / (Zx + Zy)
        return np.round(Zeq, 4) == np.round(Z1, 4)

    def at_bus_4(Ia1, Ia2, Ia0):
        # fault resistance
        ZFault = act_R_fault / (base_kv3**2 / base_mva)
        Z1 = (Vf - Va1) / Ia1
        Z2 = -(Va2 / Ia2)
        Z0 = -(Va0 / Ia0)
        Zx = complex(0, XG1) + complex(0, XT11) + ZLine + complex(0, XT21)
        Zy = ZLoad
        Zeq = Zx * Zy / (Zx + Zy)
        return np.round(Zeq, 4) == np.round(Z1, 4)
    
    # Check which case the parameter Z1 fits into
    if at_bus_1(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 1"
    elif at_bus_2(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 2"
    elif at_bus_3(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 3"
    elif at_bus_4(Ia1, Ia2, Ia0):
        return "Fault occurs at bus 4"
    else:
        return "Unknown case"

def assign_data(file_data, min, max):
    unassigned_data = extract_data_from_sheet(file_data, 'Sheet1', min, max)
    return unassigned_data

# Usage
file_path = "four-bus-power-system.xlsx"
sheet_name = "Sheet1"
system_data = extract_data_from_sheet(file_path, sheet_name, 2, 2)
bus_num = SLG_fault([[1.9734, -31.547], [1.9734, -151.547], [1.9734, 88.453]], 
[[0.8073, -31.547], [0.8073, -151.547], [0.8073, 88.453]], system_data)
print(bus_num)


