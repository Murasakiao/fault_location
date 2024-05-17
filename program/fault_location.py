from openpyxl import load_workbook # type: ignore
import math
import cmath

def extract_data_from_sheet(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data

def filter_data_by_condition(data, condition_column_index, condition_value, data_need):
    filtered_data = []
    for row in data:
        if row[condition_column_index] == condition_value:
            filtered_data.extend(row)
            time = filtered_data[0]
            filtered_data = filtered_data[1:-1]
    if data_need == 'time':
        return time
    elif data_need == 'fault_data':
        return filtered_data

def assign_values(data_array):
    values = []
    # values.append(data_array[0:2])
    # values.append(data_array[2:4])
    # values.append(data_array[4:6])
    # values.append(data_array[6:8])
    # values.append(data_array[8:10])
    # values.append(data_array[10:12])
    # values.append(data_array[12:14])
    # values.append(data_array[14:16])
    for i in range(0, len(data_array), 2):
        values.append(data_array[i:i+2])
    return values

def calc_phase_to_phase_fault_location(fault_data):
    converted_data = [[item[0], math.radians(item[1])] for item in fault_data]
    Va = converted_data[0]
    Vb = converted_data[1]
    Vc = converted_data[2]
    Ia = converted_data[3]
    Ib = converted_data[4]
    Ic = converted_data[5]
    In = converted_data[6] 
    Z1line = converted_data[7]
    Z2line = converted_data[8]
    Z0line = converted_data[9]
    BaseValues = fault_data[10]
    # perform calculations
    Vab = cmath.rect(Va[0], Va[1]) - cmath.rect(Vb[0], Vb[1])
    #Vab = [abs(Vab), math.degrees(cmath.phase(Vab))]
    Iab = cmath.rect(Ia[0], Ia[1]) - cmath.rect(Ib[0], Ib[1]) 
    #Iab = [abs(Iab), math.degrees(cmath.phase(Iab))]
    Zab = (Vab / Iab)*(BaseValues[0] / BaseValues[1])
    Z1line = cmath.rect(Z1line[0], Z1line[1])
    fault_location = abs(Zab / Z1line) * fault_data[11][0]
   # return [abs(Zab / Zline), math.degrees(cmath.phase(Zab / Zline))]
    return fault_location 

def calc_three_phase_fault_location(fault_data):
    converted_data = [[item[0], math.radians(item[1])] for item in fault_data]
    Va = converted_data[0]
    Vb = converted_data[1]
    Vc = converted_data[2]
    Ia = converted_data[3]
    Ib = converted_data[4]
    Ic = converted_data[5]
    In = converted_data[6] 
    Z1line = converted_data[7]
    Z2line = converted_data[8]
    Z0line = converted_data[9]
    BaseValues = fault_data[10]
    #positive sequence components
    a = cmath.rect(1, math.radians(120))
    V1 = (cmath.rect(Va[0], Va[1]) + cmath.rect(Vb[0], Vb[1])*a + cmath.rect(Vc[0], Vc[1])*a**2) / 3
    #V1 = [abs(V1)*fault_data[8][0], math.degrees(cmath.phase(V1))]
    I1 = (cmath.rect(Ia[0], Ia[1]) + cmath.rect(Ib[0], Ib[1])*a + cmath.rect(Ic[0], Ic[1])*a**2) / 3
    #I1 = [abs(I1)*fault_data[8][1], math.degrees(cmath.phase(I1))]
    Z1f = (V1 / I1) * (BaseValues[0] / BaseValues[1])
    #Z1f = [abs(Z1f), math.degrees(cmath.phase(Z1f))]
    Z1line = cmath.rect(Z1line[0], Z1line[1])
    fault_location = (Z1f / Z1line)*fault_data[11][0]
    #return [abs(fault_location), math.degrees(cmath.phase(fault_location))]
    return abs(fault_location)

def calc_phase_to_ground_fault_location(fault_data):
    converted_data = [[item[0], math.radians(item[1])] for item in fault_data]
    Va = converted_data[0]
    Vb = converted_data[1]
    Vc = converted_data[2]
    Ia = converted_data[3]
    Ib = converted_data[4]
    Ic = converted_data[5]
    In = converted_data[6] 
    Z1line = converted_data[7]
    Z2line = converted_data[8]
    Z0line = converted_data[9]
    BaseValues = fault_data[10]
    voltage_magnitudes = [Va[0], Vb[0], Vc[0]]
    current_magnitudes = [Ia[0], Ib[0], Ic[0]]
    faulted_phase_voltage = converted_data[voltage_magnitudes.index(min(voltage_magnitudes))]
    faulted_phase_current = converted_data[current_magnitudes.index(max(current_magnitudes)) + 3]
    #first calculate ko
    k = (cmath.rect(Z0line[0], Z0line[1]) - cmath.rect(Z1line[0], Z1line[1])) / (3 * cmath.rect(Z1line[0], Z1line[1]))
    #get the positive sequence of the fault impedance Zf
    Zf = (cmath.rect(faulted_phase_voltage[0], faulted_phase_voltage[1])) / (cmath.rect(faulted_phase_current[0], faulted_phase_current[1]) + 
                                                                           k * cmath.rect(In[0], In[1]))
    #fault location per unit 
    fault_location_pu = Zf / cmath.rect(Z1line[0], Z1line[1])
    #fault location 
    fault_location = fault_location_pu * (BaseValues[0] / BaseValues[1]) * fault_data[11][0]
    return abs(fault_location)

# Example usage
file_path = '03-2024-3PF_FAULT.xlsx'
sheet_name = 'Sheet1'

# Extract data from the sheet
data = extract_data_from_sheet(file_path, sheet_name)

# Filter data based on a condition
filtered_data = filter_data_by_condition(data, -1, 'yes', 'fault_data')

# Print the filtered data
print(filtered_data)
assigned_values = assign_values(filtered_data)
print(assigned_values)

fault_location = calc_three_phase_fault_location(assigned_values)
print(fault_location)
