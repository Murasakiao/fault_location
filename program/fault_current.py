from openpyxl import load_workbook # type: ignore
import numpy as np # type: ignore
import math
import cmath

def extract_data_from_sheet(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []
    all_data = []
    for row in sheet.iter_rows(min_row=2, max_row=2, values_only=True):
        data.append(row)
    for row in data:
        all_data.extend(row)
    return all_data

def calculate_3_phase_fault_current_(system_data, bus_num):
    per_unit_values = []
    # assign base values
    Vf = 1
    base_mva = system_data[5]
    base_kv1 = system_data[1]
    base_kv2 = system_data[7]
    base_kv3 = system_data[15]
    # assign actual values
    act_mva_gen = system_data[0]
    act_kv_gen = system_data[1]
    act_mva_T1 = system_data[5]
    act_kv1_T1 = system_data[6]
    act_kv2_T1 = system_data[7]
    act_mva_T2 = system_data[13]
    act_kv1_T2 = system_data[14]
    act_kv2_T2 = system_data[15]
    act_mw_load = system_data[19]
    act_mvar_load = system_data[20]
    act_R_fault = system_data[21]
    act_X_fault = system_data[22]
    act_z_ground = system_data[23]
    # assign per unit values 
    # generator
    XG1 = system_data[2] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG2 = system_data[3] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG0 = system_data[4] * (base_mva / act_mva_gen) * (act_kv2_T1 / base_kv2)**2
    # transformer 1
    XT11 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT12 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT10 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    # transmission line
    ZLine = complex(system_data[11], system_data[12]) / (base_kv2**2 / base_mva)
    # transformer 2
    XT21 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT22 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT20 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    # load
    RLoad = Vf**2 * (base_mva / act_mw_load)
    # load current
    ILoad = act_mw_load / base_mva

    if bus_num == 1:
        # fault resistance
        ZFault = act_R_fault / (base_kv1**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XG1)
        Zy = complex(0, XT11) + ZLine + complex(0, XT21) + RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = 0
        Z0 = 0
    elif bus_num == 2:
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XG1) + complex(0, XT11)
        Zy = ZLine + complex(0, XT21) + RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = 0
        Z0 = 0
    elif bus_num == 3:
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XT11) + complex(0, XG1) + ZLine
        Zy = complex(0, XT21) + RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = 0
        Z0 = 0
    elif bus_num == 4:
        # fault resistance
        ZFault = act_R_fault / (base_kv3**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XT11) + complex(0, XG1) + ZLine + complex(0, XT21)
        Zy = RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = 0
        Z0 = 0

    # sequence current 
    Ia1 = Vf / (Z1 + ZFault)
    Ia2 = 0
    Ia0 = 0
    # fault currents
    a = complex(-0.5, math.sqrt(3)/2)
    A = np.array([[1, 1, 1], [1, a**2, a], [1, a, a**2]])
    Ia012 = np.array([Ia0, Ia1, Ia2])
    Iabc = np.dot(A, Ia012)
    Ia = Iabc[0]
    Ib = Iabc[1]
    Ic = Iabc[2]
    # fault voltages
    Va = Ia * ZFault
    Vb = a**2 * Va
    Vc = a * Va
    # convert to polar and degrees
    Ia = [abs(Ia), math.degrees(cmath.phase(Ia))]
    Ib = [abs(Ib), math.degrees(cmath.phase(Ib))]
    Ic = [abs(Ic), math.degrees(cmath.phase(Ic))]
    # convert to polar and degrees
    Va = [abs(Va), math.degrees(cmath.phase(Va))]
    Vb = [abs(Vb), math.degrees(cmath.phase(Vb))]
    Vc = [abs(Vc), math.degrees(cmath.phase(Vc))]
    return [Ia, Ib, Ic]

def calculate_SLG_fault_current_(system_data, bus_num):
    per_unit_values = []
    # assign base values
    Vf = 1
    base_mva = system_data[5]
    base_kv1 = system_data[1]
    base_kv2 = system_data[7]
    base_kv3 = system_data[15]
    # assign actual values
    act_mva_gen = system_data[0]
    act_kv_gen = system_data[1]
    act_mva_T1 = system_data[5]
    act_kv1_T1 = system_data[6]
    act_kv2_T1 = system_data[7]
    act_mva_T2 = system_data[13]
    act_kv1_T2 = system_data[14]
    act_kv2_T2 = system_data[15]
    act_mw_load = system_data[19]
    act_mvar_load = system_data[20]
    act_R_fault = system_data[21]
    act_X_fault = system_data[22]
    act_z_ground = system_data[23]
    # assign per unit values 
    # generator
    XG1 = system_data[2] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG2 = system_data[3] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG0 = system_data[4] * (base_mva / act_mva_gen) * (act_kv2_T1 / base_kv2)**2
    # transformer 1
    XT11 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT12 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT10 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    # transmission line
    ZLine = complex(system_data[11], system_data[12]) / (base_kv2**2 / base_mva)
    # transformer 2
    XT21 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT22 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT20 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    # load
    RLoad = Vf**2 * (base_mva / act_mw_load)
    # fault resistance
    ZFault = act_R_fault / (base_kv2**2 / base_mva)
    # load current
    ILoad = act_mw_load / base_mva
    
    # bus logic
    if bus_num == 1:
        # fault resistance
        ZFault = act_R_fault / (base_kv1**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XG1)
        Zy = complex(0, XT11) + ZLine + complex(0, XT21) + RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = (Zx*Zy) / (Zx + Zy)
        Z0 = complex(0, XT12) + ZLine
    elif bus_num == 2:
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XG1) + complex(0, XT11)
        Zy = ZLine + complex(0, XT21) + RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = (Zx*Zy) / (Zx + Zy)
        Z0 = complex(0, XT12) + ZLine
    elif bus_num == 3:
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XT11) + complex(0, XG1) + ZLine
        Zy = complex(0, XT21) + RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = (Zx*Zy) / (Zx + Zy)
        Z0 = complex(0, XT12) + ZLine
    elif bus_num == 4:
        # fault resistance
        ZFault = act_R_fault / (base_kv3**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XT11) + complex(0, XG1) + ZLine + complex(0, XT21)
        Zy = RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = (Zx*Zy) / (Zx + Zy)
        Z0 = complex(0, XT12) + ZLine

    # current sequence 
    Ia1 = Vf / (Z1 + Z2 + Z0 + 3 * ZFault)
    Ia2 = Ia1
    Ia0 = Ia1
    # fault currents
    a = complex(-0.5, math.sqrt(3)/2)
    A = np.array([[1, 1, 1], [1, a**2, a], [1, a, a**2]])
    Ia012 = np.array([Ia0, Ia1, Ia2])
    Iabc = np.dot(A, Ia012)
    Ia = Iabc[0]
    Ib = Iabc[1]
    Ic = Iabc[2]
    # voltage sequence
    Va1 = Vf - Ia1 * Z1
    Va2 = -Ia2 * Z2
    Va0 = -Ia2 * Z0
    # fault voltages 
    Va012 = np.array([Va0, Va1, Va2])
    Vabc = np.dot(A, Va012)
    Va = Vabc[0]
    Vb = Vabc[1]
    Vc = Vabc[2]
    # convert to polar and degrees
    Ia = [abs(Ia), math.degrees(cmath.phase(Ia))]
    Ib = [abs(Ib), math.degrees(cmath.phase(Ib))]
    Ic = [abs(Ic), math.degrees(cmath.phase(Ic))]
    # convert to polar and degrees
    Va = [abs(Va), math.degrees(cmath.phase(Va))]
    Vb = [abs(Vb), math.degrees(cmath.phase(Vb))]
    Vc = [abs(Vc), math.degrees(cmath.phase(Vc))]

    return [Ia, Ib, Ic]

def calculate_LL_fault_current_(system_data, bus_num):
    per_unit_values = []
    # assign base values
    Vf = 1
    base_mva = system_data[5]
    base_kv1 = system_data[1]
    base_kv2 = system_data[7]
    base_kv3 = system_data[15]
    # assign actual values
    act_mva_gen = system_data[0]
    act_kv_gen = system_data[1]
    act_mva_T1 = system_data[5]
    act_kv1_T1 = system_data[6]
    act_kv2_T1 = system_data[7]
    act_mva_T2 = system_data[13]
    act_kv1_T2 = system_data[14]
    act_kv2_T2 = system_data[15]
    act_mw_load = system_data[19]
    act_mvar_load = system_data[20]
    act_R_fault = system_data[21]
    act_X_fault = system_data[22]
    act_z_ground = system_data[23]
    # assign per unit values 
    # generator
    XG1 = system_data[2] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG2 = system_data[3] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG0 = system_data[4] * (base_mva / act_mva_gen) * (act_kv2_T1 / base_kv2)**2
    # transformer 1
    XT11 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT12 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT10 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    # transmission line
    ZLine = complex(system_data[11], system_data[12]) / (base_kv2**2 / base_mva)
    # transformer 2
    XT21 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT22 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT20 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    # load
    RLoad = Vf**2 * (base_mva / act_mw_load)
    # fault resistance
    ZFault = act_R_fault / (base_kv2**2 / base_mva)
    # load current
    ILoad = act_mw_load / base_mva

    # bus logic
    if bus_num == 1:
        # fault resistance
        ZFault = act_R_fault / (base_kv1**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XG1)
        Zy = complex(0, XT11) + ZLine + complex(0, XT21) + RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = (Zx*Zy) / (Zx + Zy)
        Z0 = 0
    elif bus_num == 2:
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XG1) + complex(0, XT11)
        Zy = ZLine + complex(0, XT21) + RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = (Zx*Zy) / (Zx + Zy)
        Z0 = complex(0, XT12) + ZLine
    elif bus_num == 3:
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XT11) + complex(0, XG1) + ZLine
        Zy = complex(0, XT21) + RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = (Zx*Zy) / (Zx + Zy)
        Z0 = complex(0, XT12) + ZLine
    elif bus_num == 4:
        # fault resistance
        ZFault = act_R_fault / (base_kv3**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XT11) + complex(0, XG1) + ZLine + complex(0, XT21)
        Zy = RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = (Zx*Zy) / (Zx + Zy)
        Z0 = complex(0, XT12) + ZLine

    # sequence current 
    Ia1 = Vf / (Z1 + Z2 + ZFault)
    Ia2 = -Ia1
    Ia0 = 0
    # fault currents
    a = complex(-0.5, math.sqrt(3)/2)
    A = np.array([[1, 1, 1], [1, a**2, a], [1, a, a**2]])
    Ia012 = np.array([Ia0, Ia1, Ia2])
    Iabc = np.dot(A, Ia012)
    Ia = Iabc[0]
    Ib = Iabc[1]
    Ic = Iabc[2]
    # sequence voltages
    Va0 = 0
    Va1 = Vf - Ia1 * Z1
    Va2 = -Ia2 * Z2
    # fault voltages 
    Va012 = np.array([Va0, Va1, Va2])
    Vabc = np.dot(A, Va012)
    Va = Vabc[0]
    Vb = Vabc[1]
    Vc = Vabc[2]
    # convert to polar and degrees
    Ia = [abs(Ia), math.degrees(cmath.phase(Ia))]
    Ib = [abs(Ib), math.degrees(cmath.phase(Ib))]
    Ic = [abs(Ic), math.degrees(cmath.phase(Ic))]
    # convert to polar and degrees
    Va = [abs(Va), math.degrees(cmath.phase(Va))]
    Vb = [abs(Vb), math.degrees(cmath.phase(Vb))]
    Vc = [abs(Vc), math.degrees(cmath.phase(Vc))]
    return [Ia, Ib, Ic]

def calculate_DLG_fault_current_(system_data, bus_num):
    per_unit_values = []
    # assign base values
    Vf = 1
    base_mva = system_data[5]
    base_kv1 = system_data[1]
    base_kv2 = system_data[7]
    base_kv3 = system_data[15]
    # assign actual values
    act_mva_gen = system_data[0]
    act_kv_gen = system_data[1]
    act_mva_T1 = system_data[5]
    act_kv1_T1 = system_data[6]
    act_kv2_T1 = system_data[7]
    act_mva_T2 = system_data[13]
    act_kv1_T2 = system_data[14]
    act_kv2_T2 = system_data[15]
    act_mw_load = system_data[19]
    act_mvar_load = system_data[20]
    act_R_fault = system_data[21]
    act_X_fault = system_data[22]
    act_z_ground = system_data[23]
    # assign per unit values 
    # generator
    XG1 = system_data[2] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG2 = system_data[3] * (base_mva / act_mva_gen) * (act_kv_gen / base_kv1)**2
    XG0 = system_data[4] * (base_mva / act_mva_gen) * (act_kv2_T1 / base_kv2)**2
    # transformer 1
    XT11 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT12 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    XT10 = system_data[8] * (base_mva / act_mva_T1) * (act_kv2_T1 / base_kv2)**2
    # transmission line
    ZLine = complex(system_data[11], system_data[12]) / (base_kv2**2 / base_mva)
    # transformer 2
    XT21 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT22 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    XT20 = system_data[16] * (base_mva / act_mva_T2) * (act_kv2_T2 / base_kv3)**2
    # load
    RLoad = Vf**2 * (base_mva / act_mw_load)
    # load current
    ILoad = act_mw_load / base_mva

    # bus logic
    if bus_num == 1:
        # fault resistance
        ZFault = act_R_fault / (base_kv1**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XG1)
        Zy = complex(0, XT11) + ZLine + complex(0, XT21) + RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = (Zx*Zy) / (Zx + Zy)
        Z0 = complex(0, XT10) + ZLine
        ZG = act_z_ground / (base_kv2**2 / base_mva)
        Z1f = Z1 + ZFault
        Z2f = Z2 + ZFault
        Z0f3g = Z0 + ZFault + 3 * ZG
        Z20fg = Z2 + Z0 + 2 * ZFault + 3 * ZG
    elif bus_num == 2:
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XG1) + complex(0, XT11)
        Zy = ZLine + complex(0, XT21) + RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = (Zx*Zy) / (Zx + Zy)
        Z0 = complex(0, XT10) + ZLine
        ZG = act_z_ground / (base_kv2**2 / base_mva)
        Z1f = Z1 + ZFault
        Z2f = Z2 + ZFault
        Z0f3g = Z0 + ZFault + 3 * ZG
        Z20fg = Z2 + Z0 + 2 * ZFault + 3 * ZG
    elif bus_num == 3:
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XT11) + complex(0, XG1) + ZLine
        Zy = complex(0, XT21) + RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = (Zx*Zy) / (Zx + Zy)
        Z0 = complex(0, XT10) + ZLine
        ZG = act_z_ground / (base_kv2**2 / base_mva)
        Z1f = Z1 + ZFault
        Z2f = Z2 + ZFault
        Z0f3g = Z0 + ZFault + 3 * ZG
        Z20fg = Z2 + Z0 + 2 * ZFault + 3 * ZG
    elif bus_num == 4:
        # fault resistance
        ZFault = act_R_fault / (base_kv2**2 / base_mva)
        # sequence network calculation for impedance
        Zx = complex(0, XT11) + complex(0, XG1) + ZLine + complex(0, XT21)
        Zy = RLoad
        Z1 = (Zx*Zy) / (Zx + Zy)
        Z2 = (Zx*Zy) / (Zx + Zy)
        Z0 = complex(0, XT10) + ZLine
        ZG = act_z_ground / (base_kv3**2 / base_mva)
        Z1f = Z1 + ZFault
        Z2f = Z2 + ZFault
        Z0f3g = Z0 + ZFault + 3 * ZG
        Z20fg = Z2 + Z0 + 2 * ZFault + 3 * ZG

    # sequence current  
    Ia1 = Vf / (Z1f + (Z2f * Z0f3g) / (Z2f + Z0f3g))
    Ia2 = -Ia1 * (Z0f3g / Z20fg)
    Ia0 = -Ia1 * (Z2f / Z20fg)
    # fault currents
    a = complex(-0.5, math.sqrt(3)/2)
    A = np.array([[1, 1, 1], [1, a**2, a], [1, a, a**2]])
    Ia012 = np.array([Ia0, Ia1, Ia2])
    Iabc = np.dot(A, Ia012)
    Ia = Iabc[0]
    Ib = Iabc[1]
    Ic = Iabc[2]
    # sequence voltages
    Va1 = Vf - Ia1 * Z1
    Va2 = -Ia2 * Z2
    Va0 = -Ia0 * Z0
    # fault voltages 
    Va012 = np.array([Va0, Va1, Va2])
    Vabc = np.dot(A, Va012)
    Va = Vabc[0]
    Vb = Vabc[1]
    Vc = Vabc[2]
    # convert to polar and degrees
    Ia = [abs(Ia), math.degrees(cmath.phase(Ia))]
    Ib = [abs(Ib), math.degrees(cmath.phase(Ib))]
    Ic = [abs(Ic), math.degrees(cmath.phase(Ic))]
    # convert to polar and degrees
    Va = [abs(Va), math.degrees(cmath.phase(Va))]
    Vb = [abs(Vb), math.degrees(cmath.phase(Vb))]
    Vc = [abs(Vc), math.degrees(cmath.phase(Vc))]
    
    return [Ia, Ib, Ic]

# Usage
file_path = "four-bus-power-system.xlsx"
sheet_name = "Sheet1"
system_data = extract_data_from_sheet(file_path, sheet_name)

# solve for per unit values
fault_current = calculate_DLG_fault_current_(system_data, 4)
print(fault_current)
